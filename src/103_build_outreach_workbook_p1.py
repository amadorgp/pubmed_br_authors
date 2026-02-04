"""
103_build_outreach_workbook_p1.py

Purpose
-------
Create an outreach-ready Excel workbook (CRM-like) from the final lead export:
data/processed/leads_p1_with_email.csv

Output
------
data/processed/leads_p1_with_email_outreach.xlsx

Features
--------
- Adds operational columns: outreach_status, email_verified, preferred_channel, dates, owner, notes
- Adds derived fields: email_domain
- Data validation dropdowns for consistent workflow
- Filters + freeze panes + basic formatting
"""

from __future__ import annotations

from datetime import datetime
from pathlib import Path

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.table import Table, TableStyleInfo


PROJECT_ROOT = Path(__file__).resolve().parents[1]
IN_CSV = PROJECT_ROOT / "data" / "processed" / "leads_p1_with_email.csv"
OUT_XLSX = PROJECT_ROOT / "data" / "processed" / "leads_p1_with_email_outreach.xlsx"


STATUS_OPTIONS = [
    "Not Started",
    "Queued",
    "Contacted",
    "Follow-up Scheduled",
    "Replied",
    "Bounced",
    "Wrong Person",
    "Not a Fit",
    "Do Not Contact",
    "Closed",
]

EMAIL_VERIFIED_OPTIONS = ["Unverified", "Verified", "Invalid"]
CHANNEL_OPTIONS = ["Email", "LinkedIn", "Phone", "Other"]


def email_domain(email: object) -> str:
    if not isinstance(email, str) or "@" not in email:
        return ""
    return email.split("@", 1)[1].strip().lower()


def autosize_columns(ws, max_width: int = 60) -> None:
    # Simple autosize by max string length per column (capped)
    for col_idx in range(1, ws.max_column + 1):
        max_len = 0
        for row_idx in range(1, ws.max_row + 1):
            v = ws.cell(row=row_idx, column=col_idx).value
            if v is None:
                continue
            s = str(v)
            if len(s) > max_len:
                max_len = len(s)
        width = min(max(10, max_len + 2), max_width)
        ws.column_dimensions[get_column_letter(col_idx)].width = width


def add_dropdown(ws, col_letter: str, start_row: int, end_row: int, options: list[str]) -> None:
    # Excel list validation: needs a quoted comma-separated string
    formula = '"' + ",".join(options) + '"'
    dv = DataValidation(type="list", formula1=formula, allow_blank=True)
    ws.add_data_validation(dv)
    dv.add(f"{col_letter}{start_row}:{col_letter}{end_row}")


def main() -> None:
    if not IN_CSV.exists():
        raise FileNotFoundError(f"Missing input: {IN_CSV}")

    df = pd.read_csv(IN_CSV, low_memory=False)

    # Ensure a few expected columns exist (don't hard fail on missing optional ones)
    for col in ["author_cluster_id", "email_example"]:
        if col not in df.columns:
            raise ValueError(f"Input CSV missing required column: {col}")

    # Derived field
    df["email_domain"] = df["email_example"].apply(email_domain)

    # Operational columns (blank/defaults)
    df["outreach_status"] = "Not Started"
    df["email_verified"] = "Unverified"
    df["preferred_channel"] = "Email"
    df["date_first_contact"] = ""
    df["date_last_followup"] = ""
    df["next_action_date"] = ""
    df["owner"] = ""
    df["notes"] = ""

    # Reorder columns: keep existing lead fields first, then derived/ops
    ops_cols = [
        "email_domain",
        "outreach_status",
        "email_verified",
        "preferred_channel",
        "date_first_contact",
        "date_last_followup",
        "next_action_date",
        "owner",
        "notes",
    ]
    base_cols = [c for c in df.columns if c not in ops_cols]
    df = df[base_cols + ops_cols].copy()

    # Build workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Leads_P1_Email" # type: ignore

    # Metadata sheet (lightweight)
    meta = wb.create_sheet("Meta")
    meta["A1"] = "Generated at"
    meta["B1"] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    meta["A2"] = "Source"
    meta["B2"] = str(IN_CSV)
    meta["A3"] = "Rows"
    meta["B3"] = int(df.shape[0])
    meta["A4"] = "Columns"
    meta["B4"] = int(df.shape[1])
    meta["A6"] = "Notes"
    meta["B6"] = "Operational workbook for outreach. Emails are treated as contact channels; verify as needed."
    for cell in ["A1", "A2", "A3", "A4", "A6"]:
        meta[cell].font = Font(bold=True)

    # Write header
    header_font = Font(bold=True)
    for col_idx, col_name in enumerate(df.columns, start=1):
        cell = ws.cell(row=1, column=col_idx, value=col_name) # type: ignore
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # Write data
    for r_idx, row in enumerate(df.itertuples(index=False), start=2):
        for c_idx, value in enumerate(row, start=1):
            ws.cell(row=r_idx, column=c_idx, value=value) # type: ignore

    # Freeze header row
    ws.freeze_panes = "A2" # type: ignore

    # Add filters via Excel table
    end_row = ws.max_row # type: ignore
    end_col = ws.max_column # type: ignore
    table_ref = f"A1:{get_column_letter(end_col)}{end_row}"
    tab = Table(displayName="LeadsP1Email", ref=table_ref)
    tab.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium9",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )
    ws.add_table(tab) # type: ignore

    # Dropdown validations (apply to all rows)
    # Find column letters by header names
    col_map = {ws.cell(row=1, column=i).value: get_column_letter(i) for i in range(1, end_col + 1)} # type: ignore

    start_row = 2
    if end_row >= start_row:
        add_dropdown(ws, col_map["outreach_status"], start_row, end_row, STATUS_OPTIONS)
        add_dropdown(ws, col_map["email_verified"], start_row, end_row, EMAIL_VERIFIED_OPTIONS)
        add_dropdown(ws, col_map["preferred_channel"], start_row, end_row, CHANNEL_OPTIONS)

    # Align a few operational columns
    for name in ["outreach_status", "email_verified", "preferred_channel", "owner"]:
        letter = col_map.get(name)
        if not letter:
            continue
        for r in range(2, end_row + 1):
            ws[f"{letter}{r}"].alignment = Alignment(horizontal="left", vertical="center") # type: ignore

    # Autosize widths
    autosize_columns(ws)

    # Ensure output directory exists
    OUT_XLSX.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUT_XLSX)

    print("OK: 103_build_outreach_workbook_p1")
    print(f"Input:  {IN_CSV} | shape={df.shape}")
    print(f"Output: {OUT_XLSX}")


if __name__ == "__main__":
    main()
